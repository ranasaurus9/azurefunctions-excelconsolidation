import logging
import io
import xlrd
import pandas as pd
import openpyxl
import datetime
import azure.functions as func
from azure.storage.blob import BlobServiceClient, BlobClient, ContainerClient
from io import BytesIO
import xlsxwriter
import os
from os import listdir
from os.path import isfile, join
import pyodbc


def otemp(rsheet,  ws1, zone, country, city1, brewery, dte, cleaned_inputblob_name):
    pr = 11
    pc = 1
    while (pr < 63):
        process = ws1.cell_value(pr, pc)
        if process in 'Packaging Lines.':
            pr = pr + 1
        while True:
            line = ws1.cell_value(pr+1,pc+1)
            if line == "":
                break
            mylist = []
            mylist.append(cleaned_inputblob_name)
            mylist.append(zone)
            mylist.append(country)
            mylist.append(city1)
            mylist.append(brewery)
            mylist.append(dte)
            mylist.append(process)
            mylist.append(line)
            for i in range(pc+2,21):
                if i == 15:
                    continue
                mylist.append(ws1.cell_value(pr+1,i))
            my_list = [mylist]
            df = pd.DataFrame(my_list, columns= rsheet.columns )
            rsheet = rsheet.append(df)
            pr = pr + 1
        pr = pr + 2
    return rsheet
    # sheetnames = ['Overview_template', 'Overview_graph', 'Brewhouse_equipment', 'Tanks', 'Mother_beer_mix', 'Extract_Losses', 'Brewhouse', 'Brewhouse_details', 'KRAUZEN', 'Fermentation',
    #               'Warm_Maturation', 'Cellar_transfers', 'Cold_Aging', 'Filtration_Preparation', 'BBT', 'BBT_Cellars', 'Glass_Lines', 'Can_Lines', 'PET_lines', 'KEG_lines', 'Special_KEG_lines']
    # df_collection = {}
    # count = 0
    # for i in sheetnames:
    #     if i == 'Overview_template':
    #         df_collection[count] = rsheet
    #         count = count + 1
    #         continue
    #     dtemp = pd.read_excel(blob_bytes, sheet_name=i)
    #     print(dtemp)
    #     df_collection[count] = dtemp
    #     count = count + 1

    # xlb = io.BytesIO()
    # writer = pd.ExcelWriter(xlb, engine='xlsxwriter')

    # for i in range(0, len(sheetnames)):
    #     df_collection[i].to_excel(
    #         writer, sheet_name=sheetnames[i], index=False)
    # writer.save()
    # xlb.seek(0)
    # blob_client.upload_blob(xlb, overwrite=True)
    logging.info("otemp complete")


def ograph(rsheet14,  ws1, zone, country, city1, brewery, dte, cleaned_inputblob_name):
    pr = 298
    pc = 52
    vr,vc = 50,3
    while pc < 86 :
        if (ws1.cell_value(pr,pc) == ""):
            pc = pc+1
            vc = vc + 1
            continue
        mylist = []
        mylist.append(cleaned_inputblob_name)
        mylist.append(zone)
        mylist.append(country)
        mylist.append(city1)
        mylist.append(brewery)
        mylist.append(dte)
        for i in range(0,9):
            mylist.append(ws1.cell_value(pr+i,pc))
        mylist.append(ws1.cell_value(vr,vc))
        my_list = [mylist]
        df = pd.DataFrame(my_list, columns= rsheet14.columns )
        rsheet14 = rsheet14.append(df)
        pc = pc + 1
        vc = vc + 1
        
    return rsheet14
    logging.info("ograph complete")


def proc1(rsheet15,  ws1, zone, country, city1, brewery, dte, cleaned_inputblob_name):
    pr = 23
    pc = 1

    while pr < 33 :
        flag = 0
        for i in range(2,25):
            if(ws1.cell_value(pr,pc+i) != 0):
                flag = 1
                break
            
        if(flag == 0):
            break
        mylist = []
        mylist.append(cleaned_inputblob_name)
        mylist.append(zone)
        mylist.append(country)
        mylist.append(city1)
        mylist.append(brewery)
        mylist.append(dte)
        for i in range(0,25):
            if i == 1:
                continue
            mylist.append(ws1.cell_value(pr,pc+i))
        my_list = [mylist]
        df = pd.DataFrame(my_list, columns= rsheet15.columns )
        rsheet15 = rsheet15.append(df)    
        pr = pr + 1
        
    return rsheet15
    logging.info("proc1 complete")


def proc2(rsheet16,  ws1, zone, country, city1, brewery, dte, cleaned_inputblob_name):
    pr = 41
    pc = 1

    while pr < 51 :
        flag = 0
        for i in range(2,27):
            if(ws1.cell_value(pr,pc+i) != 0):
                flag = 1
                break
    
        if(flag == 0):
            break
        mylist = []
        mylist.append(cleaned_inputblob_name)
        mylist.append(zone)
        mylist.append(country)
        mylist.append(city1)
        mylist.append(brewery)
        mylist.append(dte)
        for i in range(0,27):
            if i == 1:
                continue
            mylist.append(ws1.cell_value(pr,pc+i))
            
        my_list = [mylist]
        df = pd.DataFrame(my_list, columns= rsheet16.columns )
        rsheet16 = rsheet16.append(df)    
        pr = pr + 1
        
    return rsheet16
    logging.info("proc2 complete")


def mobeermix(rsheet2,  ws1, zone, country, city1, brewery, dte, cleaned_inputblob_name):
    pr = 4
    pc = 14
    if "Extract" in ws1.cell_value(9,8):
        flag = 1
    else:
        flag = 0
    while pc<104:
        if ws1.cell_value(pr,pc) == 'n.a':
            pc = pc + 1
            continue
        mylist = []
        mylist.append(cleaned_inputblob_name)
        mylist.append(zone)
        mylist.append(country)
        mylist.append(city1)
        mylist.append(brewery)
        mylist.append(dte)
        i = pr
        if flag == 1:
            while(i < 17):
                if i == 9:
                    i = i+1
                    continue
                mylist.append(ws1.cell_value(i,pc))
                i = i+1
        else:
            while(i < 16):
                mylist.append(ws1.cell_value(i,pc))
                i = i+1
        pc = pc + 1
        my_list = [mylist]
        df = pd.DataFrame(my_list, columns= rsheet2.columns )
        rsheet2 = rsheet2.append(df)
    
    return rsheet2
    logging.info("OK")


def exlosses(rsheet3,  ws1, zone, country, city1, brewery, dte, cleaned_inputblob_name):
    pr = 6
    pc = 4
    while pc < 104:
        if ws1.cell_value(pr,pc) == 'n.a':
            pc = pc + 1
            continue
        mylist = []
        mylist.append(cleaned_inputblob_name)
        mylist.append(zone)
        mylist.append(country)
        mylist.append(city1)
        mylist.append(brewery)
        mylist.append(dte)
        i = pr
        while(i < 32):
            if(i in [9,10,11,22]):
                i = i + 1
                continue
            mylist.append(ws1.cell_value(i,pc))
            i = i+1
        pc = pc + 1
        my_list = [mylist]
        df = pd.DataFrame(my_list, columns= rsheet3.columns )
        rsheet3 = rsheet3.append(df)

    return rsheet3
    logging.info("OK")


def brewh(rsheet4,  ws1, zone, country, city1, brewery, dte, cleaned_inputblob_name):
    mbeer = 0
    i = 28
    j = 4
    while j < 104:
        if(ws1.cell_value(i,j) == 'n.a'):
            j = j + 1
            continue
        mbeer = mbeer + 1
        j = j + 1
    pr = 37
    pc = 1
    pcc = 4
    cbeer = 0
    while (pr < 267):
        brewl = ws1.cell_value(pr,pc)
        pcc = 4
        prr = pr + 1
        while pcc < j:
            if ws1.cell_value(prr,pcc) == 'n.a':
                pcc = pcc + 1
                continue
            mylist = []
            mylist.append(cleaned_inputblob_name)
            mylist.append(zone)
            mylist.append(country)
            mylist.append(city1)
            mylist.append(brewery)
            mylist.append(dte)
            mylist.append(brewl)
            for i in range(0,15):
                mylist.append(ws1.cell_value(prr+i,pcc))

            pcc = pcc + 1
            cbeer = cbeer + 1
            my_list = [mylist]
            df = pd.DataFrame(my_list, columns= rsheet4.columns )
            rsheet4 = rsheet4.append(df)
        pr = pr + 23
        
    return rsheet4
    logging.info("OK")


def brewdet(rsheet5,  ws1, zone, country, city1, brewery, dte, cleaned_inputblob_name):


    mbeer = 0
    i = 38
    j = 4
    while j < 104:
        if(ws1.cell_value(i,j) == 'n.a'):
            j = j + 1
            continue
        mbeer = mbeer + 1
        j = j + 1
    
    if "MOTHER" in ws1.cell_value(44,1):
        pr = 78
        cons = 241
    else:
        pr = 45
        cons = 208
    
    pc = 1
    pcc = 4
    cbeer = 0
    while (pr < cons):
        brewl = ws1.cell_value(pr,pc)
        pcc = 4
        prr = pr + 1
        while pcc < j:
            if ws1.cell_value(prr,pcc) == 'n.a':
                pcc = pcc + 1
                continue
            mylist = []
            mylist.append(cleaned_inputblob_name)
            mylist.append(zone)
            mylist.append(country)
            mylist.append(city1)
            mylist.append(brewery)
            mylist.append(dte)
            mylist.append(brewl)
            for i in range(1,16):
                mylist.append(ws1.cell_value(prr+i,pcc))

            pcc = pcc + 1
            cbeer = cbeer + 1
            my_list = [mylist]
            df = pd.DataFrame(my_list, columns= rsheet5.columns )
            rsheet5 = rsheet5.append(df)
        pr = pr + 18
        
    return rsheet5
    logging.info("OK")


def krauzy(rsheet6,  ws1, zone, country, city1, brewery, dte, cleaned_inputblob_name):
    i,j = 30,4
    while j < 104:
        if(ws1.cell_value(i,j) == 'n.a'):
            j = j + 1
            continue
        j = j+1
        
    pr = 40
    pc = 1

    while pr < 350:
        tanki = ws1.cell_value(pr,pc)
        prr = pr + 1
        pcc = pc + 3
        while pcc < j:
            if ws1.cell_value(prr,pcc) == 'n.a':
                pcc = pcc + 1
                continue
            mylist = []
            mylist.append(cleaned_inputblob_name)
            mylist.append(zone)
            mylist.append(country)
            mylist.append(city1)
            mylist.append(brewery)
            mylist.append(dte)
            mylist.append(tanki)

            for i in range(0,23):
                mylist.append(ws1.cell_value(prr+i,pcc))

            pcc = pcc + 1  
            my_list = [mylist]
            df = pd.DataFrame(my_list, columns= rsheet6.columns )
            rsheet6 = rsheet6.append(df)
        pr = pr + 31
        
    return rsheet6
    logging.info("OK")


def ferment(rsheet7,  ws1, zone, country, city1, brewery, dte, cleaned_inputblob_name):

    i,j = 32,4
    while j < 104:
        if(ws1.cell_value(i,j) == 'n.a'):
            j = j + 1
            continue
            
        j = j+1
    
    if "MOTHER" in ws1.cell_value(40,1):
        flag = 1
        pr = 49
        cons = 365
        cons2 = 35
    else:
        flag = 0
        pr = 41
        cons = 348
        cons2 = 34

    pc = 1
    while pr < cons:
        tanki = ws1.cell_value(pr,pc)
        prr = pr + 1
        pcc = pc + 3
        while pcc < j:
            if ws1.cell_value(prr,pcc) == 'n.a':
                pcc = pcc + 1
                continue
            mylist = []
            mylist.append(cleaned_inputblob_name)
            mylist.append(zone)
            mylist.append(country)
            mylist.append(city1)
            mylist.append(brewery)
            mylist.append(dte)
            mylist.append(tanki)
            if flag == 1:
                for i in range(0,28):
                    if i in [2,3]:
                        continue
                    mylist.append(ws1.cell_value(prr+i,pcc))
            else:   
                for i in range(0,26):
                    mylist.append(ws1.cell_value(prr+i,pcc))
                
            my_list = [mylist]
            df = pd.DataFrame(my_list, columns= rsheet7.columns )
            rsheet7 = rsheet7.append(df)
            pcc = pcc + 1
        pr = pr + cons2
        
    return rsheet7
    logging.info("OK")


def maturat(rsheet8,  ws1, zone, country, city1, brewery, dte, cleaned_inputblob_name):

    i,j = 30,4
    while j < 104:
        if(ws1.cell_value(i,j) == 'n.a'):
            j = j + 1
            continue
        j = j+1
        
    pr = 39
    pc = 1

    while pr < 339:
        tanki = ws1.cell_value(pr,pc)
        prr = pr + 1
        pcc = pc + 3

        while pcc < j:
            if ws1.cell_value(prr,pcc) == 'n.a':
                pcc = pcc + 1
                continue
            mylist = []
            mylist.append(cleaned_inputblob_name)
            mylist.append(zone)
            mylist.append(country)
            mylist.append(city1)
            mylist.append(brewery)
            mylist.append(dte)
            mylist.append(tanki)
            for i in range(0,22):
                mylist.append(ws1.cell_value(prr+i,pcc))
                
            my_list = [mylist]
            df = pd.DataFrame(my_list, columns= rsheet8.columns )
            rsheet8 = rsheet8.append(df)
            pcc = pcc + 1
        pr = pr + 30
    
    return rsheet8
    logging.info("OK")


def cellar(rsheet9,  ws1, zone, country, city1, brewery, dte, cleaned_inputblob_name):
    i,j = 29,4
    while j < 104:
        if(ws1.cell_value(i,j) == 'n.a'):
            j = j +1
            continue
        j = j+1
    pr = 38
    pc = 1
    
    while pr < 268:
        tanki = ws1.cell_value(pr,pc)
        prr = pr + 1
        pcc = pc + 3

        while pcc < j:
            if ws1.cell_value(prr,pcc) == 'n.a':
                pcc = pcc + 1
                continue
            mylist = []
            mylist.append(cleaned_inputblob_name)
            mylist.append(zone)
            mylist.append(country)
            mylist.append(city1)
            mylist.append(brewery)
            mylist.append(dte)
            mylist.append(tanki)
            for i in range(0,15):
                mylist.append(ws1.cell_value(prr+i,pcc))
                
            my_list = [mylist]
            df = pd.DataFrame(my_list, columns= rsheet9.columns )
            rsheet9 = rsheet9.append(df)
            pcc = pcc + 1
        pr = pr + 23
        
    return rsheet9
    logging.info("OK")


def coldag(rsheet10,  ws1, zone, country, city1, brewery, dte, cleaned_inputblob_name):

    i,j = 30,4
    while j < 104:
        if(ws1.cell_value(i,j) == 'n.a'):
            j = j + 1
            continue
            
        j = j+1
        
    pr = 39
    pc = 1

    while pr < 339:
        tanki = ws1.cell_value(pr,pc)
        prr = pr + 1
        pcc = pc + 3

        while pcc < j:
            if ws1.cell_value(prr,pcc) == 'n.a':
                pcc = pcc + 1
                continue
            mylist = []
            mylist.append(cleaned_inputblob_name)
            mylist.append(zone)
            mylist.append(country)
            mylist.append(city1)
            mylist.append(brewery)
            mylist.append(dte)
            mylist.append(tanki)
            for i in range(0,22):
                mylist.append(ws1.cell_value(prr+i,pcc))
                
            my_list = [mylist]
            df = pd.DataFrame(my_list, columns= rsheet10.columns )
            rsheet10 = rsheet10.append(df)
            pcc = pcc + 1
        pr = pr + 30
    
    return rsheet10
    logging.info("OK")


def filtprep(rsheet11,  ws1, zone, country, city1, brewery, dte, cleaned_inputblob_name):
    i,j,fline = 7,4,0
    while ws1.cell_value(i,j) != 'n.a':
        fline = fline + 1
        j = j + 1

    
    i,j,mbeer = 32,4,0
    while j < 104:
        if ws1.cell_value(i,j) == 'n.a':
            j = j + 1
            continue
        mbeer = mbeer + 1
        j = j + 1
        
    if "Chiller" in ws1.cell_value(39,1):
        pr = 43
    else:
        pr = 41

    pc = 1
    itera = 0
    while itera < fline:
        tanki = ws1.cell_value(pr,pc)
        prr = pr + 1
        pcc = pc + 3
        
        while ((pcc - 4) < mbeer) :
            if ws1.cell_value(prr,pcc) == 'n.a':
                pcc = pcc + 1
                continue
            mylist = []
            mylist.append(cleaned_inputblob_name)
            mylist.append(zone)
            mylist.append(country)
            mylist.append(city1)
            mylist.append(brewery)
            mylist.append(dte)
            mylist.append(tanki)
            for i in range(0,15):
                mylist.append(ws1.cell_value(prr+i,pcc))
            my_list = [mylist]
            df = pd.DataFrame(my_list, columns= rsheet11.columns )
            rsheet11 = rsheet11.append(df)
            pcc = pcc + 1
        pr = pr + 23
        itera = itera + 1
    
    return rsheet11
    logging.info("OK")


def bbt1(rsheet12,  ws1, zone, country, city1, brewery, dte, cleaned_inputblob_name):
    pr = 15
    pc = 4
    while pc < 104:
        if (ws1.cell_value(pr,pc) == 'n.a'):
            pc = pc + 1
            continue
        mylist = []
        mylist.append(cleaned_inputblob_name)
        mylist.append(zone)
        mylist.append(country)
        mylist.append(city1)
        mylist.append(brewery)
        mylist.append(dte)
        for i in range(0,23):
            if (pr+i) in [20,24,29,34,36]:
                continue
            mylist.append(ws1.cell_value(pr+i,pc))
            
        my_list = [mylist]
        df = pd.DataFrame(my_list, columns= rsheet12.columns )
        rsheet12 = rsheet12.append(df)
        pc = pc + 1
        
    return rsheet12
    logging.info("bbt1")


def bbt2(rsheet13,  ws1, zone, country, city1, brewery, dte, cleaned_inputblob_name):
    i,j,tno = 7,4,0
    while(ws1.cell_value(i,j) != 'n.a') and (j<24):
        tno = tno + 1
        j = j + 1
        
    pr = 40
    pc = 4
    itera = 0
    while itera < tno :        
        mylist = []
        mylist.append(cleaned_inputblob_name)
        mylist.append(zone)
        mylist.append(country)
        mylist.append(city1)
        mylist.append(brewery)
        mylist.append(dte)
        for i in range(0,26):
            if (pr+i) in [41,42]:
                continue
            mylist.append(ws1.cell_value(pr+i,pc))

        my_list = [mylist]
        df = pd.DataFrame(my_list, columns= rsheet13.columns )
        rsheet13 = rsheet13.append(df)
        pc = pc + 1
        itera = itera + 1
        
    return rsheet13
    logging.info("bb2")


def gbottle(rsheet17,  ws1, zone, country, city1, brewery, dte, cleaned_inputblob_name):
    lr = 5
    lj = 1

    while True:
        prr, pcc = lr+7, lj + 1
        flag2 = 0
        for i in range(1, 41):
            if(ws1.cell_value(prr, pcc+i) != 0) and (ws1.cell_value(prr, pcc+i) != ""):
                flag2 = 1
                break
            if(ws1.cell_value(prr + 51, pcc+i) != 0) and (ws1.cell_value(prr + 51, pcc+i) != ""):
                flag2 = 1
                break
        if flag2 == 0:
            break

        for jk in range(0, 2):
            pr = prr
            pc = pcc
            while True:
                flag = 0
                for i in range(1, 41):
                    if(ws1.cell_value(pr, pc+i) != 0) and (ws1.cell_value(pr, pc+i) != ""):
                        flag = 1
                        break
                if flag == 0:
                    break

                mylist = []
                mylist.append(cleaned_inputblob_name)
                mylist.append(zone)
                mylist.append(country)
                mylist.append(city1)
                mylist.append(brewery)
                mylist.append(dte)
                mylist.append(ws1.cell_value(lr, lj))
                mylist.append(ws1.cell_value(prr, pc))
                for i in range(1, 41):
                    if i == 37:
                        continue
                    mylist.append(ws1.cell_value(pr, pc+i))
                pr = pr + 1
                my_list = [mylist]
                df = pd.DataFrame(my_list, columns=rsheet17.columns)
                rsheet17 = rsheet17.append(df)
            prr = prr + 51
        lr = lr + 140

    return rsheet17
    logging.info("OK")


def canl(rsheet18,  ws1, zone, country, city1, brewery, dte, cleaned_inputblob_name):
    lr = 5
    lj = 1

    while True:
        prr, pcc = lr+7, lj
        flag2 = 0
        for i in range(2, 42):
            if(ws1.cell_value(prr, pcc+i) != 0) and (ws1.cell_value(prr, pcc+i) != ""):
                flag2 = 1
                break
        if flag2 == 0:
            break

        pr = prr
        pc = pcc
        while True:
            flag = 0
            for i in range(2, 42):
                if(ws1.cell_value(pr, pc+i) != 0) and (ws1.cell_value(pr, pc+i) != ""):
                    flag = 1
                    break
            if flag == 0:
                break

            mylist = []
            mylist.append(cleaned_inputblob_name)
            mylist.append(zone)
            mylist.append(country)
            mylist.append(city1)
            mylist.append(brewery)
            mylist.append(dte)
            mylist.append(ws1.cell_value(lr, lj))
            mylist.append(ws1.cell_value(prr, pc))
            for i in range(2, 42):
                if i == 38:
                    continue
                mylist.append(ws1.cell_value(pr, pc+i))
            my_list = [mylist]
            df = pd.DataFrame(my_list, columns=rsheet18.columns)
            rsheet18 = rsheet18.append(df)
            pr = pr + 1
        lr = lr + 100

    return rsheet18
    logging.info("can1")


def petl(rsheet19,  ws1, zone, country, city1, brewery, dte, cleaned_inputblob_name):
    lr = 5
    lj = 1

    while True:
        prr, pcc = lr+7, lj
        flag2 = 0
        for i in range(2, 42):
            if(ws1.cell_value(prr, pcc+i) != 0) and (ws1.cell_value(prr, pcc+i) != ""):
                flag2 = 1
                break
        if flag2 == 0:
            break

        pr = prr
        pc = pcc
        while True:
            flag = 0
            for i in range(2, 42):
                if(ws1.cell_value(pr, pc+i) != 0) and (ws1.cell_value(pr, pc+i) != ""):
                    flag = 1
                    break
            if flag == 0:
                break

            mylist = []
            mylist.append(cleaned_inputblob_name)
            mylist.append(zone)
            mylist.append(country)
            mylist.append(city1)
            mylist.append(brewery)
            mylist.append(dte)
            mylist.append(ws1.cell_value(lr, lj))
            mylist.append(ws1.cell_value(prr, pc))
            for i in range(2, 42):
                if i == 38:
                    continue
                mylist.append(ws1.cell_value(pr, pc+i))
            my_list = [mylist]
            df = pd.DataFrame(my_list, columns=rsheet19.columns)
            rsheet19 = rsheet19.append(df)
            pr = pr + 1
        lr = lr + 100

    return rsheet19
    logging.info("pet1")


def keg1(rsheet20,  ws1, zone, country, city1, brewery, dte, cleaned_inputblob_name):
    lr = 5
    lj = 1

    while True:
        prr, pcc = lr+7, lj
        flag2 = 0
        for i in range(2, 43):
            if(ws1.cell_value(prr, pcc+i) != 0) and (ws1.cell_value(prr, pcc+i) != ""):
                flag2 = 1
                break
        if flag2 == 0:
            break

        pr = prr
        pc = pcc
        while True:
            flag = 0
            for i in range(2, 43):
                if(ws1.cell_value(pr, pc+i) != 0) and (ws1.cell_value(pr, pc+i) != ""):
                    flag = 1
                    break
            if flag == 0:
                break

            mylist = []
            mylist.append(cleaned_inputblob_name)
            mylist.append(zone)
            mylist.append(country)
            mylist.append(city1)
            mylist.append(brewery)
            mylist.append(dte)
            mylist.append(ws1.cell_value(lr, lj))

            for i in range(2, 43):
                if i == 39:
                    continue
                mylist.append(ws1.cell_value(pr, pc+i))

            my_list = [mylist]
            df = pd.DataFrame(my_list, columns=rsheet20.columns)
            rsheet20 = rsheet20.append(df)
            pr = pr + 1
        lr = lr + 100

    return rsheet20
    logging.info("keg1")


def skeg1(rsheet21,  ws1, zone, country, city1, brewery, dte, cleaned_inputblob_name):
    lr = 5
    lj = 1

    while True:
        prr, pcc = lr+7, lj
        flag2 = 0
        for i in range(2, 43):
            if(ws1.cell_value(prr, pcc+i) != 0) and (ws1.cell_value(prr, pcc+i) != ""):
                flag2 = 1
                break
        if flag2 == 0:
            break

        pr = prr
        pc = pcc
        while True:
            flag = 0
            for i in range(2, 43):
                if(ws1.cell_value(pr, pc+i) != 0) and (ws1.cell_value(pr, pc+i) != ""):
                    flag = 1
                    break
            if flag == 0:
                break

            mylist = []
            mylist.append(cleaned_inputblob_name)
            mylist.append(zone)
            mylist.append(country)
            mylist.append(city1)
            mylist.append(brewery)
            mylist.append(dte)
            mylist.append(ws1.cell_value(lr, lj))

            for i in range(2, 43):
                if i == 39:
                    continue
                mylist.append(ws1.cell_value(pr, pc+i))

            my_list = [mylist]
            df = pd.DataFrame(my_list, columns=rsheet21.columns)
            rsheet21 = rsheet21.append(df)
            pr = pr + 1
        lr = lr + 100
    return rsheet21
    logging.info("skeg")


def pdown(rsheet22, cleaned_inputblob_name, s1, zone, country, city1, brewery, dte):
    x, y = 5, 1
    ws1 = s1.sheet_by_name("4.1 - GLASS BOTTLE LINES ")
    while x < 2105:
        flag1 = 0
        pr, pc = x + 112, y + 28
        for i in range(0, 4):
            if(ws1.cell_value(pr+i, pc) != ""):
                flag1 = 1
                break
        if flag1 == 0:
            x = x + 140
            continue
        mylist = []
        mylist.append(cleaned_inputblob_name)
        mylist.append(zone)
        mylist.append(country)
        mylist.append(city1)
        mylist.append(brewery)
        mylist.append(dte)
        mylist.append(ws1.cell_value(x, y))
        mylist.append("Glass")
        mylist.append(ws1.cell_value(pr, pc))
        mylist.append(ws1.cell_value(pr+1, pc))
        mylist.append(ws1.cell_value(pr+2, pc))
        mylist.append(ws1.cell_value(pr+3, pc))

        my_list = [mylist]
        df = pd.DataFrame(my_list, columns=rsheet22.columns)
        rsheet22 = rsheet22.append(df)
        x = x + 140

    x, y = 5, 1
    ws1 = s1.sheet_by_name("4.2 - CAN LINES")
    while x < 1005:
        flag1 = 0
        pr, pc = x + 60, y + 28
        for i in range(0, 4):
            if(ws1.cell_value(pr+i, pc) != ""):
                flag1 = 1
                break
        if flag1 == 0:
            x = x + 100
            continue
        mylist = []
        mylist.append(cleaned_inputblob_name)
        mylist.append(zone)
        mylist.append(country)
        mylist.append(city1)
        mylist.append(brewery)
        mylist.append(dte)
        mylist.append(ws1.cell_value(x, y))
        mylist.append("Cans")
        mylist.append(ws1.cell_value(pr, pc))
        mylist.append(ws1.cell_value(pr+1, pc))
        mylist.append(ws1.cell_value(pr+2, pc))
        mylist.append(ws1.cell_value(pr+3, pc))

        my_list = [mylist]
        df = pd.DataFrame(my_list, columns=rsheet22.columns)
        rsheet22 = rsheet22.append(df)
        x = x + 100

    x, y = 5, 1
    ws1 = s1.sheet_by_name("4.3 - PET LINES")
    while x < 1005:
        flag1 = 0
        pr, pc = x + 60, y + 28
        for i in range(0, 4):
            if(ws1.cell_value(pr+i, pc) != ""):
                flag1 = 1
                break
        if flag1 == 0:
            x = x + 100
            continue
        mylist = []
        mylist.append(cleaned_inputblob_name)
        mylist.append(zone)
        mylist.append(country)
        mylist.append(city1)
        mylist.append(brewery)
        mylist.append(dte)
        mylist.append(ws1.cell_value(x, y))
        mylist.append("PET")
        mylist.append(ws1.cell_value(pr, pc))
        mylist.append(ws1.cell_value(pr+1, pc))
        mylist.append(ws1.cell_value(pr+2, pc))
        mylist.append(ws1.cell_value(pr+3, pc))

        my_list = [mylist]
        df = pd.DataFrame(my_list, columns=rsheet22.columns)
        rsheet22 = rsheet22.append(df)
        x = x + 100

    x, y = 5, 1
    ws1 = s1.sheet_by_name("4.4 - KEG LINES")
    while x < 705:
        flag1 = 0
        pr, pc = x + 60, y + 29
        for i in range(0, 4):
            if(ws1.cell_value(pr+i, pc) != ""):
                flag1 = 1
                break
        if flag1 == 0:
            x = x + 100
            continue
        mylist = []
        mylist.append(cleaned_inputblob_name)
        mylist.append(zone)
        mylist.append(country)
        mylist.append(city1)
        mylist.append(brewery)
        mylist.append(dte)
        mylist.append(ws1.cell_value(x, y))
        mylist.append("Kegs")
        mylist.append(ws1.cell_value(pr, pc))
        mylist.append(ws1.cell_value(pr+1, pc))
        mylist.append(ws1.cell_value(pr+2, pc))
        mylist.append(ws1.cell_value(pr+3, pc))

        my_list = [mylist]
        df = pd.DataFrame(my_list, columns=rsheet22.columns)
        rsheet22 = rsheet22.append(df)
        x = x + 100

    x, y = 5, 1
    ws1 = s1.sheet_by_name("4.5 - SPECIAL KEG LINES")
    while x < 705:
        flag1 = 0
        pr, pc = x + 60, y + 29
        for i in range(0, 4):
            if(ws1.cell_value(pr+i, pc) != ""):
                flag1 = 1
                break
        if flag1 == 0:
            x = x + 100
            continue
        mylist = []
        mylist.append(cleaned_inputblob_name)
        mylist.append(zone)
        mylist.append(country)
        mylist.append(city1)
        mylist.append(brewery)
        mylist.append(dte)
        mylist.append(ws1.cell_value(x, y))
        mylist.append("Special Kegs")
        mylist.append(ws1.cell_value(pr, pc))
        mylist.append(ws1.cell_value(pr+1, pc))
        mylist.append(ws1.cell_value(pr+2, pc))
        mylist.append(ws1.cell_value(pr+3, pc))

        my_list = [mylist]
        df = pd.DataFrame(my_list, columns=rsheet22.columns)
        rsheet22 = rsheet22.append(df)
        x = x + 100
    return rsheet22


async def main(inputblob: func.InputStream):
    logging.info(f"Python blob trigger function processed blob \n"
                 f"Name: {inputblob.name}\n")
    s1 = xlrd.open_workbook(file_contents=inputblob.read())
    ws1 = s1.sheet_by_name('1.0 - MENU')
    zone = ws1.cell_value(9, 3)
    country = ws1.cell_value(11, 3)
    city1 = ws1.cell_value(13, 3)
    brewery = ws1.cell_value(15, 3)

    cleaned_inputblob_name = inputblob.name
    cleaned_inputblob_name = cleaned_inputblob_name[8:]
    resarray = {"AFR": "results-AFR.xlsx", "NAZ": "results-NAZ.xlsx", "EUR": "results-EUR.xlsx",
                "APA": "results-APAC.xlsx", "MAZ": "results-MAZ.xlsx", "SAZ": "results-SAZ.xlsx"}
    dte = cleaned_inputblob_name[cleaned_inputblob_name.find(
        "20"):cleaned_inputblob_name.find(".xlsx")]

# #   connect_str = 'DefaultEndpointsProtocol=https;AccountName=storageaccountcapaca500;AccountKey=BbOiUl8IqLOE7HtGsw349vi3RLzZXzKlGfaNf5aR7fXGF8AYOwYWCyNGGFgFquIoVvpo0rm+nCx5iwPkvXN82A==;EndpointSuffix=core.windows.net'

#     connect_str = "DefaultEndpointsProtocol=https;AccountName=capacityfilessagbdev;AccountKey=LXjTpW6qRZo25w8eLLxOa/Oae4qeziprr4GTcFDgCAwKy4VPy35h8xL1yJLqoY/VJ3HjSQsXRdqUiEJeCeqpqQ==;EndpointSuffix=core.windows.net"
#     # Create the BlobServiceClient object which will be used to create a container client
#     blob_service_client = BlobServiceClient.from_connection_string(connect_str)
# #   container_name = 'result'
#     container_name = 'results'

# #   container_client = blob_service_client.create_container(container_name)
#     blob_client = blob_service_client.get_blob_client(
#         container=container_name, blob=resarray[cleaned_inputblob_name[0:3]])
#     blob = blob_client.download_blob()
#     blob_bytes = blob_client.download_blob().content_as_bytes()

    # connection to SQL DB
    server = 'tcp:capacityplanningserver.database.windows.net,1433'
    database = 'Capacity DB'
    username = 'SRV_Global_CapacityPlanning@ab-inbev.com'
    driver = '{ODBC Driver 17 for SQL Server}'
    password = 'Xf^A2wNg6}+rfum!'

    with pyodbc.connect('DRIVER='+driver+';SERVER='+server+';DATABASE='+database+';UID='+username+';PWD='+password+';Authentication=ActiveDirectoryPassword') as conn:
        with conn.cursor() as cursor:
            print("Connected")

            heade = ["Filename", "Zone", "Country","City","Brewery","Date","Process","Line","Bottle_Neck","Equipment_Type","Equipment_Nbr","Production_Peak_Month_percent","Sales_Seasonal_Factor","Production_Seasonal_Factor",	"Process_Losses_or_LEF_percent","Dilution_Factor_or_GLY_percent",	"Engineering_Capacity_Week_khl",	"Engineering_Capacity_Month_khl",	"Engineering_Capacity_Year_khl",	"Year_Operational_Capacity_khl",	"Budget_Volume_Year_khl",	"Budget_Volume_Production_Peak_Month_khl",	"Capacity_Utilization_Engineering_Peak_Month_percent",	"Capacity_Utilization_Engineering_Year_percent",	"Capacity_Utilization_Operational_Year_percent"]

            head14 = ["Filename", "Zone", "Country","City","Brewery","Date", "Process", "Week_Volume_in_Peak_Month_khl", "Peak_Month_Volume_khl", "Year_Volume_khl", "Year_Volume_khl_1", "Week_Engineering_Capacity_khl", "Month_Engineering_Capacity_khl", "Year_Engineering_Capacity_khl",	"Year_Operational_Capacity_khl", "Vessels_or_Lines"]

            head15 = ["Filename", "Zone", "Country","City","Brewery","Date", "Brewline","Main_Supplier","Construction_Year","Adjunct_cooker_nbr", "Adjunct_cooker_gross_volume", "Conversion_vessel_nbr",	"Conversion_vessel_gross_volume", "Mash_filtration_Technology", "Mash_filtration_nbr", "Filtration_surface_m2", "Ideal_load_kg_malt_equivalent_m2", "Weak_Wort_Tank",	 "Buffer_Tank_nbr",	"Boiling_kettle_technology", "Boiling_kettle_nbr", "Boiling_kettle_gross_volume", "Inline_preheating",	 "Trub_separation_Technology", "Trub_separation_nbr", "Trub_separation_gross_volume", "Wort_cooler_nbr", "Dosing_point_Maximum_flow_hl_h", "Syrup_dosing", "Comments"]

            head16 = ["Filename", "Zone", "Country","City","Brewery","Date", "Tank_Group", "KRAUZEN_Main_Supplier",	"KRAUZEN_Year",	"KRAUZEN_Type",	"KRAUZEN_Tanks", "KRAUZEN_Gross_Vol_per_tank_hl	","FERMENTATION_Main_Supplier",	"FERMENTATION_Year", "FERMENTATION_Type	", "FERMENTATION_Tanks", "FERMENTATION_Gross_Vol_per_tank_hl",	"WARM_MATURATION_Main_Supplier", "WARM_MATURATION_Year", "WARM_MATURATION_Type", "WARM_MATURATION_Tanks", "WARM_MATURATION_Gross_Volume_hl" , "COLD_AGING_Main_Supplier",	"COLD_AGING_Year", "COLD_AGING_Type", "COLD_AGING_Tanks	", "COLD_AGING_Gross_Vol_per_tank_hl", "BBT_Main_Supplier", "BBT_Year", "BBT_Type", "BBT_Tanks", "BBT_Gross_Vol_per_tank_hl"]

            head2 = ["Filename", "Zone", "Country","City","Brewery","Date","Mother_Beers_activated","Mother_Beers_name","Average_Sales_Beer_Gravity_FG_Plato","Input_Average_Sales_Beer_Gravity_FG_Plato","Average_Sales_Beer_Gravity_FG_kg_per_hl","Diluted_Mother_Beer_Mix_Average_Sales_Gravity_percent_Volume","Input_Diluted_Mother_Beer_Mix_Average_Sales_Gravity_percent_Volume","Cold_Wort_Gravity_Inlet_Fermentation_Plato","Cold_Wort_Gravity_Inlet_Fermentation_kg_per_hl","Total_High_Gravity_Mother_beer_used_hl","Total_equivalent_FP_produced_with_Mother_beer_hl","Mother_beer_weighted_dilution_factor_v_v"]

            head3 = ["Filename", "Zone", "Country","City","Brewery","Date","Mother_Beers_activated","Mother_Beers_name","Diluted_Mother_Beer_Mix_Average_Sales_Gravity_percent_Volume","Raw_Material_Handling","Brewhouse",	"Fermentation","Warm_Maturation","Cellar_Transfer","Cold_aging","Filtration","BBT","Packaging","Total_Extract_Losses","Total_Extract_losses_cumulative","Extract_losses_from_Raw_material_outlet_to_final_product_percent","Extract_losses_from_brewhouse_outlet_to_final_product_percent",	"Extract_losses_from_FermentationTank_Outlet_to_final_product","Extract_losses_from_Warm_Maturation_Outlet_to_final_product","Extract_losses_from_Cellar_Transfer_outlet_to_final_product",	"Extract_losses_from_Cold_Aging_Outlet_to_final_product","Extract_losses_from_beer_Filtration_outlet_to_final_product","Extract_losses_from_Beer_outlet_BBT_to_final_product"]

            head4 = ["Filename", "Zone", "Country","City","Brewery","Date", "Brewline","Mother_Beers_activated","Mother_Beers_name","Cold_Wort_Gravity_Inlet_Fermentation_Plato","	Cold_Wort_Gravity_Inlet_Fermentation_kg_per_hl","Brewlength_Cold_Wort_fermentation_inlet_hl","Brewline_takt_time_hours","Input_Brewline_takt_time_hours","Theoretical_Cold_Wort_Output_Rate_hl_HG_per_h",	"One_Week_168_h_maximum_HG_cold_wort_volume_inlet_Fermentation_hl_HG",	"Cold_Wort_Mix_percent_Volume_of_Cold_Wort_HG_percent_Volume","Occupation_Time_percent",	"Occupation_time_h","Extract_losses_from_cold_wort_to_final_product_percent	","Average_Dilution_Factor",	"One_Week_168_h_maximum_sales_beer_volume_produced_hl_FG"]

            head5 = ["Filename", "Zone", "Country","City","Brewery","Date", "Brewline","Mother_Beers_activated","Mother_Beers_name", "Input_Adjunct_cooker_cycle_time","Input_Conversion_vessel_cycle_time",	"Input_Mash_filtration_cycle_time","Input_Boiling_kettle_cycle_time","Input_Trub_separation_cycle_time","Input_Wort_cooler_cycle_time","Adjunct_cooker_takt_time","Conversion_vessel_takt_time",	"Mash_filtration_takt_time","Boiling_kettle_takt_time","Trub_separation_takt_time","Wort_cooler_takt_time","Brewline_takt_time"]

            head6 = ["Filename", "Zone", "Country","City","Brewery","Date", "Tank_Group", "Mother_Beers_activated","Mother_Beers_name","Krauzening_ratio_percent_Volume","Filling_rate_percent", "Working_volume_per_tank_hl", "Working_volume_per_all_tanks_of_group_hl","Beer_Gravity_Tank_Outlet_kg_per_hl", "Filling_Time_hours", "Process_Time_for_Krauzen_Preparation_hours","Emptying_Time_hours", "Tank_CIP_Cycle_Time_hours", "Accepted_Downtime_hours", "Phasing_hours", "Scheduling_hours"," Manual_connections_hours", "Other_hours","Total_Cycle_Time_hours", "Corresponding_Output_Rate_hl_HG_per_h", "One_Week_168_h_maximum_HG_Beer_Volume_Tanks_Group_Outlet_hl_HG", "Mother_Beer_Mix_Tank_Outlet_percent_Volume", "Extract_losses_from_Tank_Outlet_to_final_product_percent", "Average_Dilution_Factor_from_Tank_Outlet_to_Final_Product" ,"One_Week_168_h_maximum_sales_beer_volume_produced_hl"]

            head7 = ["Filename", "Zone", "Country","City","Brewery","Date", "Tank_Group", "Mother_Beers_activated","Mother_Beers_name","Filling_rate_percent","Working_volume_per_tank_hl", "Working_volume_per_all_tanks_of_group_hl", "Beer_Gravity_Tank_Outlet_kg_per_hl", "Filling_Time_hours", "Process_Time_for_Extract_Consumption_Fermentation_hours", "Process_Time_for_diacetyl_Reduction_Maturation_hours", "Process_Time_for_Cooling_prior_to_Emptying_hours", "Process_Time_for_Cold_Aging_in_case_of_unitank_process_hours", "Process_Time_hours", "Emptying_Time_hours", "Tank_CIP_Cycle_Time_hours", "Accepted_Downtime_hours", "Phasing_hours", "Scheduling_hours", "Manual_connections_hours", "Other_hours", "Total_Cycle_Time_hours", "Corresponding_Output_Rate_hl_HG_per_h", "One_Week_168_h_maximum_HG_Beer_Volume_Tanks_Group_Outlet_hl_HG", "Mother_Beer_Mix_Tank_Outlet_percent_Volume", "Extract_losses_from_Tank_Outlet_to_final_product_percent", "Average_Dilution_Factor_from_Tank_Outlet_to_Final_Product", "One_Week_168_h_maximum_sales_beer_volume_produced_hl"]

            head8 = ["Filename", "Zone", "Country","City","Brewery","Date", "Tank_Group", "Mother_Beers_activated","Mother_Beers_name","Filling_rate_percent", "Working_volume_per_tank_hl", "Working_volume_per_all_tanks_of_group_hl", "Beer_Gravity_Warm_Maturation_Outlet_kg_per_hl", "Filling_Time_hours", "Process_Time_Warm_Maturation_hours", "Emptying_Time_hours","Tank_CIP_Cycle_Time_hours","Accepted_Downtime_hours"," Phasing_hours ","Scheduling_hours", "Manual_connections_hours", "Other_hours", "Total_Cycle_Time_hours", "Corresponding_Output_Rate_hl_HG_per_h", "One_Week_168_h_maximum_HG_Beer_Volume_Tank_Group_Outlet","Mother_Beer_Mix_Warm_Maturation_Outlet_percent_volume", "Extract_losses_from_Warm_Maturation_Outlet_to_final_product_percent", "Average_Dilution_Factor_from_Warm_Maturation_Outlet_to_Final_Product", "One_Week_168_h_maximum_sales_beer_volume_produced_hl"]

            head9 = ["Filename", "Zone", "Country","City","Brewery","Date", "Tank_Group", "Mother_Beers_activated","Mother_Beers_name","Centrifuging_type_during_transfer ","Chilling_during_transfer", "Chilling_capability_from_warm_maturation_temperature_Superchill", "Green_Beer_Gravity_Cellar_transfers_Outlet_plato", "Green_Beer_Gravity_Cellar_transfers_Outlet_kg_per_hl", "Transfer_Line_Output_Rate_hl_HG_per_h", "One_Week_168_h_maximum_HG_cold_wort_volume_inlet_Cold_Aging_hl_HG", "Beer_Mix_inlet_Cold_Aging_percent_volume", "Occupation_Time_percent", "Occupation_time_hours", "Extract_losses_from_Green_Beer_inlet_Cold_Aging_to_final_product_percent","Dilution_Factor", "One_Week_168_h_maximum_sales_beer_volume_produced_hl"]

            head10 = ["Filename", "Zone", "Country","City","Brewery","Date", "Tank_Group", "Mother_Beers_activated","Mother_Beers_name", "Filling_rate_percent", "Working_volume_per_tank_hl", "Working_volume_per_all_tanks_of_group_hl", "Beer_Gravity_Cold_Aging_Outlet_kg_per_hl", "Filling_Time_hours", "Process_Time_for_Cold_Aging_hours", "Emptying_Time_hours", "Tank_CIP_Cycle_Time_hours", "Accepted_Downtime_hours", "Phasing_hours", "Scheduling_hours", "Manual_connections_hours", "Other_hours", "Total_Cycle_Time_hours", "Corresponding_Output_Rate_hl_HG_per_h", "One_Week_168_h_maximum_HG_Beer_Volume_Tank_Group_Outlet_hl_HG", "Mother_Beer_Mix_Cold_Aging_Outlet_percent_volume", "Extract_losses_from_Cold_Aging_Outlet_to_final_product_percent", "Average_Dilution_Factor_from_Cold_Aging_Outlet_to_Final_Product", "One_Week_168_h_maximum_sales_beer_volume_produced_hl"]

            head11 = ["Filename", "Zone", "Country","City","Brewery","Date", "Tank_Group", "Mother_Beers_activated","Mother_Beers_name", "Centrifuge", "Filtration_type", "Stabilization_type", "Average_Beer_Gravity_Filter_Inlet_plato", "Average_Beer_Gravity_Filter_inlet_kg_per_hl", "Filter_Output_Rate_supplier_Specification_hl_per_h", "One_Week_168_h_maximum_HG_volume_BBT_hl_HG", "Beer_Mix_inlet_BBT_percent_volume", "Occupation_time_percent", "Occupation_time_hours", "Extract_losses_from_Beer_inlet_BBT_to_final_product_percent", "Average_Dilution_Factor", "One_Week_168_h_maximum_sales_beer_volume_produced_hl"]

            head12 = ["Filename", "Zone", "Country","City","Brewery","Date", "Mother_Beers_activated","Mother_Beers_name", "Average_Sales_Beer_Gravity_plato", "Average_Sales_Beer_Gravity_kg_per_hl", "Diluted_Mother_Beer_Mix_Average_Sales_Gravity", "Part_of_volume_by_passing_BBT_percent", "Part_of_volume_going_in_BBT_percent", "Mother_beer_mix_in_BBT_percent", "Part_of_beer_with_additional_dilution_just_before_filler_inlet_percent", "Beer_Gravity_at_beer_filtration_plato", "Beer_Gravity_at_beer_filtration_kg_per_hl", "Average_dilution_factor" ,"Part_of_beer_with_partial_filling_in_BBT_percent", "Corresponding_filling_rate_in_case_of_partial_volume_percent", "Corresponding_filling_rate_reduction_for_partial_volume_percent", "Extended_average_rest_time_for_small_batch_hours", "Extract_losses_from_Beer_outlet_BBT_to_final_product_percent", "Frequency_number_of_cycles_between_CIP"]

            head13 = ["Filename", "Zone", "Country","City","Brewery","Date", "Tank_Group", "Tanks_nbr", "Gross_Volume_per_Tank_hl", "Main_Supplier", "Year", "Type", "Average_Filling_Rate_percent","Net_Volume_hl", "Average_Filling_Time_hours", "Minimum_Residence_Time_hours", "Tank_Group_used_for_small_batches", "Additional_time_for_small_batch_multiple_release_hours", "Average_Emptying_Time_hours", "Accepted_Downtime_hours", "Phasing_hours" ,"Scheduling_hours", "Manual_connections_hours", "Other_please_specify_in_free_text_area_below", "Average_Cycle_Time_hours_excluding_CIP", "Real_CIP_Time_hours_per_CIP", "Average_CIP_Time_hours_lost_per_cycle", "Average_Cycle_Time_Including_CIP_hours", "Average_nbr_Cycles_per_Week", "Week_Available_Bright_Beer_Tank_Capacity_hl"]

            head17 = ["Filename", "Zone", "Country","City","Brewery","Date", "Line", "Line_Name", "SKU_Container_Type", "Bottle_Content_liters",	"Bottle_Nominal_Speed_Bph", "Can_Week_Maximum_Capacity_hl", "Can_directly_to_Transport_pack_or_pallet", "SKU_secondary_Pack_Type", "Content_Bottles_per_secondary_pack	", "Content_hl_per_secondary_pack", "Calculated_Nominal_Speed_secondary_packs_per_hrs", "Forced_Nominal_Speed_secondary_packs_per_hrs", "secondary_pack_Week_Maximum_Capacity_hl", "SKU_Transport_Pack_Type", "Content_Bottle_per_transport_pack", "Content_hl_per_transport_pack", "Calculated_Nominal_Speed_Transport_packs_per_hrs", "Forced_Nominal_Speed_Transport_packs_per_hrs", "Transport_pack_Week_Maximum_Capacity_hl", "Pallets_Packs_per_layer", "Pallet_nbr_layers", "Packs_per_pallet","hl_per_pallet",	"Calculated_Nominal_Speed_pallets_per_hrs", "Forced_Nominal_Speed_pallets_per_hrs","Pallets_Week_Maximum_Capacity_hl", "Lowest_Week_Maximum_Capacity_hl",	"LEF_percent", "Week_GLY_OAE_percent", "Calculated_MIX_Capacity", "Forced_MIX_Capacity", "Time_to_Schedule_hrs_per_year", "Week_Engineering_Capacity_khl", "Month_Engineering_Capacity_khl", "Year_Engineering_Capacity_khl", "Sales_Peak_avg_3_months_percent", "Seasonal_factor",	"Year_Operational_Capacity_khl", "Total_Budgeted_Volumes_per_SKU_type_SKU_MIX_percent", "Year_Total_Budgeted_Volumes_per_SKU_type", "P_month_Total_Budgeted_Volumes_per_SKU_type"]

            head18 = ["Filename", "Zone", "Country","City","Brewery","Date", "Line", "Line_Name", "SKU_Container_Type", "Can_Content_liters",	 "Can_Nominal_Speed_Bph", "Can_Week_Maximum_Capacity_hl", "Can_directly_to_Transport_pack_or_pallet", "SKU_secondary_Pack_Type", "Content_Cans_per_secondary_pack	", "Content_hl_per_secondary_pack", "Calculated_Nominal_Speed_secondary_packs_per_hrs", "Forced_Nominal_Speed_secondary_packs_per_hrs", "secondary_pack_Week_Maximum_Capacity_hl", "SKU_Transport_Pack_Type", "Content_Can_per_transport_pack", "Content_hl_per_transport_pack", "Calculated_Nominal_Speed_Transport_packs_per_hrs", "Forced_Nominal_Speed_Transport_packs_per_hrs", "Transport_pack_Week_Maximum_Capacity_hl", "Pallets_Cans_per_layer", "Pallet_nbr_layers", "Cans_per_pallet","hl_per_pallet",	"Calculated_Nominal_Speed_pallets_per_hrs", "Forced_Nominal_Speed_pallets_per_hrs","Pallets_Week_Maximum_Capacity_hl", "Lowest_Week_Maximum_Capacity_hl",	"LEF_percent", "Week_GLY_OAE_percent", "Calculated_MIX_Capacity", "Forced_MIX_Capacity", "Time_to_Schedule_hrs_per_year", "Week_Engineering_Capacity_khl", "Month_Engineering_Capacity_khl", "Year_Engineering_Capacity_khl", "Sales_Peak_avg_3_months_percent", "Seasonal_factor",	"Year_Operational_Capacity_khl", "Total_Budgeted_Volumes_per_SKU_type_SKU_MIX_percent", "Year_Total_Budgeted_Volumes_per_SKU_type", "P_month_Total_Budgeted_Volumes_per_SKU_type"]

            head19 = ["Filename", "Zone", "Country","City","Brewery","Date", "Line", "Line_Name", "SKU_Container_Type", "PET_Content_liters",	 "PET_Nominal_Speed_Bph", "PET_Week_Maximum_Capacity_hl", "PET_directly_to_Transport_pack_or_pallet", "SKU_secondary_Pack_Type", "Content_PETs_per_secondary_pack", "Content_hl_per_secondary_pack", "Calculated_Nominal_Speed_secondary_packs_per_hrs", "Forced_Nominal_Speed_secondary_packs_per_hrs", "secondary_pack_Week_Maximum_Capacity_hl", "SKU_Transport_Pack_Type", "Content_PET_per_transport_pack", "Content_hl_per_transport_pack", "Calculated_Nominal_Speed_Transport_packs_per_hrs", "Forced_Nominal_Speed_Transport_packs_per_hrs", "Transport_pack_Week_Maximum_Capacity_hl", "Pallets_PETs_per_layer", "Pallet_nbr_layers", "PETs_per_pallet","	hl_per_pallet",	"Calculated_Nominal_Speed_pallets_per_hrs", "Forced_Nominal_Speed_pallets_per_hrs","Pallets_Week_Maximum_Capacity_hl", "Lowest_Week_Maximum_Capacity_hl",	"LEF_percent", "Week_GLY_OAE_percent", "Calculated_MIX_Capacity", "Forced_MIX_Capacity", "Time_to_Schedule_hrs_per_year", "Week_Engineering_Capacity_khl", "Month_Engineering_Capacity_khl", "Year_Engineering_Capacity_khl", "Sales_Peak_avg_3_months_percent", "Seasonal_factor",	"Year_Operational_Capacity_khl", "Total_Budgeted_Volumes_per_SKU_type_SKU_MIX_percent", "Year_Total_Budgeted_Volumes_per_SKU_type", "P_month_Total_Budgeted_Volumes_per_SKU_type"]

            head20 = ["Filename", "Zone", "Country","City","Brewery","Date", "Line", "SKU_Keg_Type", "Keg_Content_liters",	 "Keg_Nominal_Speed_Bph", "Keg_Week_Maximum_Capacity_hl", "Keg_directly_to_Transport_pack_or_pallet", "SKU_secondary_Pack_Type", "Content_Kegs_per_secondary_pack", "Content_hl_per_secondary_pack", "Calculated_Nominal_Speed_secondary_packs_per_hrs", "Forced_Nominal_Speed_secondary_packs_per_hrs", "secondary_pack_Week_Maximum_Capacity_hl","Kegs_or_secondary_pack_directly_to__pallet", "SKU_Transport_Pack_Type", "Content_Kegs_per_transport_pack", "Content_hl_per_transport_pack", "Calculated_Nominal_Speed_Transport_packs_per_hrs", "Forced_Nominal_Speed_Transport_packs_per_hrs", "Transport_pack_Week_Maximum_Capacity_hl", "Pallets_Kegs_per_layer", "Pallet_nbr_layers", "Kegs_per_pallet","	hl_per_pallet",	"Calculated_Nominal_Speed_pallets_per_hrs", "Forced_Nominal_Speed_pallets_per_hrs","Pallets_Week_Maximum_Capacity_hl", "Lowest_Week_Maximum_Capacity_hl",	"LEF_percent", "Week_GLY_OAE_percent", "Calculated_MIX_Capacity", "Forced_MIX_Capacity", "Time_to_Schedule_hrs_per_year", "Week_Engineering_Capacity_khl", "Month_Engineering_Capacity_khl", "Year_Engineering_Capacity_khl", "Sales_Peak_avg_3_months_percent", "Seasonal_factor",	"Year_Operational_Capacity_khl", "Total_Budgeted_Volumes_per_SKU_type_SKU_MIX_percent", "Year_Total_Budgeted_Volumes_per_SKU_type", "P_month_Total_Budgeted_Volumes_per_SKU_type"]

            head21 = ["Filename", "Zone", "Country","City","Brewery","Date", "Line", "SKU_Keg_Type", "Keg_Content_liters",	 "Keg_Nominal_Speed_Bph", "Keg_Week_Maximum_Capacity_hl", "Keg_directly_to_Transport_pack_or_pallet", "SKU_secondary_Pack_Type", "Content_Kegs_per_secondary_pack", "Content_hl_per_secondary_pack", "Calculated_Nominal_Speed_secondary_packs_per_hrs", "Forced_Nominal_Speed_secondary_packs_per_hrs", "secondary_pack_Week_Maximum_Capacity_hl","Kegs_or_secondary_pack_directly_to__pallet", "SKU_Transport_Pack_Type", "Content_Kegs_per_transport_pack", "Content_hl_per_transport_pack", "Calculated_Nominal_Speed_Transport_packs_per_hrs", "Forced_Nominal_Speed_Transport_packs_per_hrs", "Transport_pack_Week_Maximum_Capacity_hl", "Pallets_Kegs_per_layer", "Pallet_nbr_layers", "Kegs_per_pallet","	hl_per_pallet",	"Calculated_Nominal_Speed_pallets_per_hrs", "Forced_Nominal_Speed_pallets_per_hrs","Pallets_Week_Maximum_Capacity_hl", "Lowest_Week_Maximum_Capacity_hl",	"LEF_percent", "Week_GLY_OAE_percent", "Calculated_MIX_Capacity", "Forced_MIX_Capacity", "Time_to_Schedule_hrs_per_year", "Week_Engineering_Capacity_khl", "Month_Engineering_Capacity_khl", "Year_Engineering_Capacity_khl", "Sales_Peak_avg_3_months_percent", "Seasonal_factor",	"Year_Operational_Capacity_khl", "Total_Budgeted_Volumes_per_SKU_type_SKU_MIX_percent", "Year_Total_Budgeted_Volumes_per_SKU_type", "P_month_Total_Budgeted_Volumes_per_SKU_type"]

            head22 = ["Filename", "Zone", "Country","City","Brewery","Date","Line","Glass","Week_Total_Time","Week_Non_Schedule_Time","Week_Schedule_Time","Downtime_for_Yearly_Overhaul"]

            df_collection = {}
            # if file name contains -1.xlsx then file_name - "-1" = original filename
            # delete data where filename = file_name
            print("1st Sheet")
            ws1 = s1.sheet_by_name('1.1 - OVERVIEW TEMPLATE')
            rsheet = pd.DataFrame(columns=heade)
            df_collection[0] = otemp(rsheet,  ws1,
                                     zone, country, city1, brewery, dte, cleaned_inputblob_name)
            df_collection[0].columns = heade
            list_of_tuples = [
                tuple(x) for x in df_collection[0].to_records(index=False)]
            cursor.fast_executemany = True
            sql_statement = "INSERT INTO Overview_template (Filename ,Zone ,Country ,City, Brewery ,Date ,Process ,Line ,Bottle_Neck ,Equipment_Type ,Equipment_Nbr ,Production_Peak_Month_percent ,Sales_Seasonal_Factor ,Production_Seasonal_Factor ,Process_Losses_or_LEF_percent ,Dilution_Factor_or_GLY_percent ,Engineering_Capacity_Week_khl ,Engineering_Capacity_Month_khl ,Engineering_Capacity_Year_khl ,Year_Operational_Capacity_khl ,Budget_Volume_Year_khl ,Budget_Volume_Production_Peak_Month_khl ,Capacity_Utilization_Engineering_Peak_Month_percent ,Capacity_Utilization_Engineering_Year_percent ,Capacity_Utilization_Operational_Year_percent) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"
            if len(list_of_tuples) > 0:
                cursor.executemany(sql_statement, list_of_tuples)

            print("1st-1 Sheet")
            ws1 = s1.sheet_by_name('1.2 - OVERVIEW GRAPH')
            rsheet14 = pd.DataFrame(columns=head14)
            df_collection[1] = ograph(rsheet14,  ws1,
                                      zone, country, city1, brewery, dte, cleaned_inputblob_name)
            df_collection[1].columns = head14
            list_of_tuples = [
                tuple(x) for x in df_collection[1].to_records(index=False)]
            cursor.fast_executemany = True
            sql_statement = "INSERT INTO Overview_graph (Filename ,Zone ,Country,City,Brewery ,Date ,Process ,Week_Volume_in_Peak_Month_khl ,Peak_Month_Volume_khl ,Year_Volume_khl ,Year_Volume_khl_1 ,Week_Engineering_Capacity_khl ,Month_Engineering_Capacity_khl ,Year_Engineering_Capacity_khl ,Year_Operational_Capacity_khl ,Vessels_or_Lines) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"
            if len(list_of_tuples) > 0:
                cursor.executemany(sql_statement, list_of_tuples)

            print("1st-2 Sheet")
            ws1 = s1.sheet_by_name('1.4 - PROCESS EQUIPMENT')
            rsheet15 = pd.DataFrame(columns=head15)
            df_collection[2] = proc1(rsheet15,  ws1,
                                     zone, country, city1, brewery, dte, cleaned_inputblob_name)
            df_collection[2].columns = head15
            list_of_tuples = [
                tuple(x) for x in df_collection[2].to_records(index=False)]
            cursor.fast_executemany = True
            sql_statement = "INSERT INTO Brewhouse_equipment (Filename , Zone , Country ,City , Brewery , Date , Brewline , Main_Supplier , Construction_Year , Adjunct_cooker_nbr , Adjunct_cooker_gross_volume , Conversion_vessel_nbr , Conversion_vessel_gross_volume , Mash_filtration_Technology , Mash_filtration_nbr , Filtration_surface_m2 , Ideal_load_kg_malt_equivalent_m2 , Weak_Wort_Tank , Buffer_Tank_nbr , Boiling_kettle_technology , Boiling_kettle_nbr , Boiling_kettle_gross_volume , Inline_preheating , Trub_separation_Technology , Trub_separation_nbr , Trub_separation_gross_volume , Wort_cooler_nbr , Dosing_point_Maximum_flow_hl_h , Syrup_dosing , Comments) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"
            if len(list_of_tuples) > 0:
                cursor.executemany(sql_statement, list_of_tuples)

            print("1st-3 Sheet")
            ws1 = s1.sheet_by_name('1.4 - PROCESS EQUIPMENT')
            rsheet16 = pd.DataFrame(columns=head16)
            df_collection[3] = proc2(rsheet16,  ws1,
                                     zone, country, city1, brewery, dte, cleaned_inputblob_name)
            df_collection[3].columns = head16
            list_of_tuples = [
                tuple(x) for x in df_collection[3].to_records(index=False)]
            cursor.fast_executemany = True
            sql_statement = "INSERT INTO Tanks ( Filename , Zone , Country , City , Brewery , Date , Tank_Group , KRAUZEN_Main_Supplier , KRAUZEN_Year , KRAUZEN_Type , KRAUZEN_Tanks , KRAUZEN_Gross_Vol_per_tank_hl , FERMENTATION_Main_Supplier , FERMENTATION_Year , FERMENTATION_Type , FERMENTATION_Tanks , FERMENTATION_Gross_Vol_per_tank_hl , WARM_MATURATION_Main_Supplier , WARM_MATURATION_Year , WARM_MATURATION_Type , WARM_MATURATION_Tanks , WARM_MATURATION_Gross_Volume_hl , COLD_AGING_Main_Supplier , COLD_AGING_Year , COLD_AGING_Type , COLD_AGING_Tanks , COLD_AGING_Gross_Vol_per_tank_hl , BBT_Main_Supplier , BBT_Year , BBT_Type , BBT_Tanks , BBT_Gross_Vol_per_tank_hl) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"
            if len(list_of_tuples) > 0:
                cursor.executemany(sql_statement, list_of_tuples)

            print("2nd Sheet")
            ws1 = s1.sheet_by_name('2.0 - MOTHER BEER MIX')
            rsheet2 = pd.DataFrame(columns=head2)
            df_collection[4] = mobeermix(rsheet2,  ws1,
                                         zone, country, city1, brewery, dte, cleaned_inputblob_name)
            df_collection[4].columns = head2
            list_of_tuples = [
                tuple(x) for x in df_collection[4].to_records(index=False)]
            cursor.fast_executemany = True
            sql_statement = "INSERT INTO Mother_beer_mix ( Filename , Zone , Country , City , Brewery , Date , Mother_Beers_activated , Mother_Beers_name , Average_Sales_Beer_Gravity_FG_Plato , Input_Average_Sales_Beer_Gravity_FG_Plato , Average_Sales_Beer_Gravity_FG_kg_per_hl , Diluted_Mother_Beer_Mix_Average_Sales_Gravity_percent_Volume , Input_Diluted_Mother_Beer_Mix_Average_Sales_Gravity_percent_Volume , Cold_Wort_Gravity_Inlet_Fermentation_Plato , Cold_Wort_Gravity_Inlet_Fermentation_kg_per_hl , Total_High_Gravity_Mother_beer_used_hl , Total_equivalent_FP_produced_with_Mother_beer_hl , Mother_beer_weighted_dilution_factor_v_v) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) "
            if len(list_of_tuples) > 0:
                cursor.executemany(sql_statement, list_of_tuples)

            print("3rd Sheet")
            ws1 = s1.sheet_by_name('2.0.1 - EXTRACT LOSSES')
            rsheet3 = pd.DataFrame(columns=head3)
            df_collection[5] = exlosses(rsheet3,  ws1,
                                        zone, country, city1, brewery, dte, cleaned_inputblob_name)
            df_collection[5].columns = head3
            list_of_tuples = [
                tuple(x) for x in df_collection[5].to_records(index=False)]
            cursor.fast_executemany = True
            sql_statement = "INSERT INTO Extract_Losses ( Filename , Zone , Country , City , Brewery , Date , Mother_Beers_activated , Mother_Beers_name , Diluted_Mother_Beer_Mix_Average_Sales_Gravity_percent_Volume , Raw_Material_Handling , Brewhouse , Fermentation , Warm_Maturation , Cellar_Transfer , Cold_aging , Filtration , BBT , Packaging , Total_Extract_Losses , Total_Extract_losses_cumulative , Extract_losses_from_Raw_material_outlet_to_final_product_percent , Extract_losses_from_brewhouse_outlet_to_final_product_percent , Extract_losses_from_FermentationTank_Outlet_to_final_product , Extract_losses_from_Warm_Maturation_Outlet_to_final_product , Extract_losses_from_Cellar_Transfer_outlet_to_final_product , Extract_losses_from_Cold_Aging_Outlet_to_final_product , Extract_losses_from_beer_Filtration_outlet_to_final_product , Extract_losses_from_Beer_outlet_BBT_to_final_product ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"
            if len(list_of_tuples) > 0:
                cursor.executemany(sql_statement, list_of_tuples)

            print("4th Sheet")
            ws1 = s1.sheet_by_name('2.2 - BREWHOUSE')
            rsheet4 = pd.DataFrame(columns=head4)
            df_collection[6] = brewh(rsheet4,  ws1,
                                     zone, country, city1, brewery, dte, cleaned_inputblob_name)
            df_collection[6].columns = head4
            list_of_tuples = [
                tuple(x) for x in df_collection[6].to_records(index=False)]
            cursor.fast_executemany = True
            sql_statement = "INSERT INTO Brewhouse ( Filename , Zone , Country , City , Brewery , Date , Brewline , Mother_Beers_activated , Mother_Beers_name , Cold_Wort_Gravity_Inlet_Fermentation_Plato , Cold_Wort_Gravity_Inlet_Fermentation_kg_per_hl , Brewlength_Cold_Wort_fermentation_inlet_hl , Brewline_takt_time_hours , Input_Brewline_takt_time_hours , Theoretical_Cold_Wort_Output_Rate_hl_HG_per_h , One_Week_168_h_maximum_HG_cold_wort_volume_inlet_Fermentation_hl_HG , Cold_Wort_Mix_percent_Volume_of_Cold_Wort_HG_percent_Volume , Occupation_Time_percent , Occupation_time_h , Extract_losses_from_cold_wort_to_final_product_percent , Average_Dilution_Factor , One_Week_168_h_maximum_sales_beer_volume_produced_hl_FG) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"
            if len(list_of_tuples) > 0:
                cursor.executemany(sql_statement, list_of_tuples)

            print("5th Sheet")
            ws1 = s1.sheet_by_name('2.2.1 - BREWHOUSE DETAILS')
            rsheet5 = pd.DataFrame(columns=head5)
            df_collection[7] = brewdet(rsheet5,  ws1,
                                       zone, country, city1, brewery, dte, cleaned_inputblob_name)
            df_collection[7].columns = head5
            list_of_tuples = [
                tuple(x) for x in df_collection[7].to_records(index=False)]
            cursor.fast_executemany = True
            sql_statement = "INSERT INTO Brewhouse_details ( Filename , Zone , Country , City , Brewery , Date , Brewline , Mother_Beers_activated , Mother_Beers_name , Input_Adjunct_cooker_cycle_time , Input_Conversion_vessel_cycle_time , Input_Mash_filtration_cycle_time , Input_Boiling_kettle_cycle_time , Input_Trub_separation_cycle_time , Input_Wort_cooler_cycle_time , Adjunct_cooker_takt_time , Conversion_vessel_takt_time , Mash_filtration_takt_time , Boiling_kettle_takt_time , Trub_separation_takt_time , Wort_cooler_takt_time , Brewline_takt_time) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"
            if len(list_of_tuples) > 0:
                cursor.executemany(sql_statement, list_of_tuples)

            print("6th Sheet")
            ws1 = s1.sheet_by_name('2.3 - KRAUZEN')
            rsheet6 = pd.DataFrame(columns=head6)
            df_collection[8] = krauzy(rsheet6,  ws1,
                                      zone, country, city1, brewery, dte, cleaned_inputblob_name)
            df_collection[8].columns = head6
            list_of_tuples = [
                tuple(x) for x in df_collection[8].to_records(index=False)]
            cursor.fast_executemany = True
            sql_statement = "INSERT INTO KRAUZEN (Filename , Zone , Country , City , Brewery , Date , Tank_Group , Mother_Beers_activated , Mother_Beers_name , Krauzening_ratio_percent_Volume , Filling_rate_percent , Working_volume_per_tank_hl , Working_volume_per_all_tanks_of_group_hl , Beer_Gravity_Tank_Outlet_kg_per_hl , Filling_Time_hours , Process_Time_for_Krauzen_Preparation_hours , Emptying_Time_hours , Tank_CIP_Cycle_Time_hours , Accepted_Downtime_hours , Phasing_hours , Scheduling_hours , Manual_connections_hours , Other_hours , Total_Cycle_Time_hours , Corresponding_Output_Rate_hl_HG_per_h , One_Week_168_h_maximum_HG_Beer_Volume_Tanks_Group_Outlet_hl_HG , Mother_Beer_Mix_Tank_Outlet_percent_Volume , Extract_losses_from_Tank_Outlet_to_final_product_percent , Average_Dilution_Factor_from_Tank_Outlet_to_Final_Product , One_Week_168_h_maximum_sales_beer_volume_produced_hl) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) "
            if len(list_of_tuples) > 0:
                cursor.executemany(sql_statement, list_of_tuples)

            print("7th Sheet")
            ws1 = s1.sheet_by_name('2.4 - FERMENTATION')
            rsheet7 = pd.DataFrame(columns=head7)
            df_collection[9] = ferment(rsheet7,  ws1,
                                       zone, country, city1, brewery, dte, cleaned_inputblob_name)
            df_collection[9].columns = head7
            list_of_tuples = [
                tuple(x) for x in df_collection[9].to_records(index=False)]
            cursor.fast_executemany = True
            sql_statement = "INSERT INTO Fermentation (Filename , Zone , Country , City , Brewery , Date , Tank_Group , Mother_Beers_activated , Mother_Beers_name , Filling_rate_percent , Working_volume_per_tank_hl , Working_volume_per_all_tanks_of_group_hl , Beer_Gravity_Tank_Outlet_kg_per_hl , Filling_Time_hours , Process_Time_for_Extract_Consumption_Fermentation_hours , Process_Time_for_diacetyl_Reduction_Maturation_hours , Process_Time_for_Cooling_prior_to_Emptying_hours , Process_Time_for_Cold_Aging_in_case_of_unitank_process_hours , Process_Time_hours , Emptying_Time_hours , Tank_CIP_Cycle_Time_hours , Accepted_Downtime_hours , Phasing_hours , Scheduling_hours , Manual_connections_hours , Other_hours , Total_Cycle_Time_hours , Corresponding_Output_Rate_hl_HG_per_h , One_Week_168_h_maximum_HG_Beer_Volume_Tanks_Group_Outlet_hl_HG , Mother_Beer_Mix_Tank_Outlet_percent_Volume , Extract_losses_from_Tank_Outlet_to_final_product_percent , Average_Dilution_Factor_from_Tank_Outlet_to_Final_Product , One_Week_168_h_maximum_sales_beer_volume_produced_hl) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"
            if len(list_of_tuples) > 0:
                cursor.executemany(sql_statement, list_of_tuples)

            print("8th Sheet")
            ws1 = s1.sheet_by_name('2.5 - WARM MATURATION')
            rsheet8 = pd.DataFrame(columns=head8)
            df_collection[10] = maturat(rsheet8,  ws1,
                                        zone, country, city1, brewery, dte, cleaned_inputblob_name)
            df_collection[10].columns = head8
            list_of_tuples = [
                tuple(x) for x in df_collection[10].to_records(index=False)]
            cursor.fast_executemany = True
            sql_statement = "INSERT INTO Warm_Maturation ( Filename , Zone , Country , City , Brewery , Date , Tank_Group , Mother_Beers_activated , Mother_Beers_name , Filling_rate_percent , Working_volume_per_tank_hl , Working_volume_per_all_tanks_of_group_hl , Beer_Gravity_Warm_Maturation_Outlet_kg_per_hl , Filling_Time_hours , Process_Time_Warm_Maturation_hours , Emptying_Time_hours , Tank_CIP_Cycle_Time_hours , Accepted_Downtime_hours , Phasing_hours , Scheduling_hours , Manual_connections_hours , Other_hours , Total_Cycle_Time_hours , Corresponding_Output_Rate_hl_HG_per_h , One_Week_168_h_maximum_HG_Beer_Volume_Tank_Group_Outlet , Mother_Beer_Mix_Warm_Maturation_Outlet_percent_volume , Extract_losses_from_Warm_Maturation_Outlet_to_final_product_percent , Average_Dilution_Factor_from_Warm_Maturation_Outlet_to_Final_Product , One_Week_168_h_maximum_sales_beer_volume_produced_hl) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"
            if len(list_of_tuples) > 0:
                cursor.executemany(sql_statement, list_of_tuples)

            print("9th Sheet")
            ws1 = s1.sheet_by_name('2.6 - CELLAR TRANSFERS')
            rsheet9 = pd.DataFrame(columns=head9)
            df_collection[11] = cellar(rsheet9,  ws1,
                                       zone, country, city1, brewery, dte, cleaned_inputblob_name)
            df_collection[11].columns = head9
            list_of_tuples = [
                tuple(x) for x in df_collection[11].to_records(index=False)]
            cursor.fast_executemany = True
            sql_statement = "INSERT INTO Cellar_transfers ( Filename , Zone , Country , City , Brewery , Date , Tank_Group , Mother_Beers_activated , Mother_Beers_name , Centrifuging_type_during_transfer , Chilling_during_transfer , Chilling_capability_from_warm_maturation_temperature_Superchill , Green_Beer_Gravity_Cellar_transfers_Outlet_plato , Green_Beer_Gravity_Cellar_transfers_Outlet_kg_per_hl , Transfer_Line_Output_Rate_hl_HG_per_h , One_Week_168_h_maximum_HG_cold_wort_volume_inlet_Cold_Aging_hl_HG , Beer_Mix_inlet_Cold_Aging_percent_volume , Occupation_Time_percent , Occupation_time_hours , Extract_losses_from_Green_Beer_inlet_Cold_Aging_to_final_product_percent , Dilution_Factor , One_Week_168_h_maximum_sales_beer_volume_produced_hl) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"
            if len(list_of_tuples) > 0:
                cursor.executemany(sql_statement, list_of_tuples)

            print("10th Sheet")
            ws1 = s1.sheet_by_name('2.7 - COLD AGING')
            rsheet10 = pd.DataFrame(columns=head10)
            df_collection[12] = coldag(rsheet10,  ws1,
                                       zone, country, city1, brewery, dte, cleaned_inputblob_name)
            df_collection[12].columns = head10
            list_of_tuples = [
                tuple(x) for x in df_collection[12].to_records(index=False)]
            cursor.fast_executemany = True
            sql_statement = "INSERT INTO Cold_Aging ( Filename , Zone , Country , City , Brewery , Date , Tank_Group , Mother_Beers_activated , Mother_Beers_name , Filling_rate_percent , Working_volume_per_tank_hl , Working_volume_per_all_tanks_of_group_hl , Beer_Gravity_Cold_Aging_Outlet_kg_per_hl , Filling_Time_hours , Process_Time_for_Cold_Aging_hours , Emptying_Time_hours , Tank_CIP_Cycle_Time_hours , Accepted_Downtime_hours , Phasing_hours , Scheduling_hours , Manual_connections_hours , Other_hours , Total_Cycle_Time_hours , Corresponding_Output_Rate_hl_HG_per_h , One_Week_168_h_maximum_HG_Beer_Volume_Tank_Group_Outlet_hl_HG , Mother_Beer_Mix_Cold_Aging_Outlet_percent_volume , Extract_losses_from_Cold_Aging_Outlet_to_final_product_percent , Average_Dilution_Factor_from_Cold_Aging_Outlet_to_Final_Product , One_Week_168_h_maximum_sales_beer_volume_produced_hl) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"
            if len(list_of_tuples) > 0:
                cursor.executemany(sql_statement, list_of_tuples)

            print("11th Sheet")
            ws1 = s1.sheet_by_name('2.8 - FILTRATION & PREPARATION')
            rsheet11 = pd.DataFrame(columns=head11)
            df_collection[13] = filtprep(rsheet11,  ws1,
                                         zone, country, city1, brewery, dte, cleaned_inputblob_name)
            df_collection[13].columns = head11
            list_of_tuples = [
                tuple(x) for x in df_collection[13].to_records(index=False)]
            cursor.fast_executemany = True
            sql_statement = "INSERT INTO Filtration_Preparation ( Filename , Zone , Country , City , Brewery , Date , Tank_Group , Mother_Beers_activated , Mother_Beers_name , Centrifuge , Filtration_type , Stabilization_type , Average_Beer_Gravity_Filter_Inlet_plato , Average_Beer_Gravity_Filter_inlet_kg_per_hl , Filter_Output_Rate_supplier_Specification_hl_per_h , One_Week_168_h_maximum_HG_volume_BBT_hl_HG , Beer_Mix_inlet_BBT_percent_volume , Occupation_time_percent , Occupation_time_hours , Extract_losses_from_Beer_inlet_BBT_to_final_product_percent , Average_Dilution_Factor , One_Week_168_h_maximum_sales_beer_volume_produced_hl) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"
            if len(list_of_tuples) > 0:
                cursor.executemany(sql_statement, list_of_tuples)

            print("12th Sheet")
            ws1 = s1.sheet_by_name('2.9 - BBT')
            rsheet12 = pd.DataFrame(columns=head12)
            df_collection[14] = bbt1(rsheet12,  ws1,
                                     zone, country, city1, brewery, dte, cleaned_inputblob_name)
            df_collection[14].columns = head12
            list_of_tuples = [
                tuple(x) for x in df_collection[14].to_records(index=False)]
            cursor.fast_executemany = True
            sql_statement = "INSERT INTO BBT (Filename , Zone , Country , City , Brewery , Date , Mother_Beers_activated , Mother_Beers_name , Average_Sales_Beer_Gravity_plato , Average_Sales_Beer_Gravity_kg_per_hl , Diluted_Mother_Beer_Mix_Average_Sales_Gravity , Part_of_volume_by_passing_BBT_percent , Part_of_volume_going_in_BBT_percent , Mother_beer_mix_in_BBT_percent , Part_of_beer_with_additional_dilution_just_before_filler_inlet_percent , Beer_Gravity_at_beer_filtration_plato , Beer_Gravity_at_beer_filtration_kg_per_hl , Average_dilution_factor , Part_of_beer_with_partial_filling_in_BBT_percent , Corresponding_filling_rate_in_case_of_partial_volume_percent , Corresponding_filling_rate_reduction_for_partial_volume_percent , Extended_average_rest_time_for_small_batch_hours , Extract_losses_from_Beer_outlet_BBT_to_final_product_percent , Frequency_number_of_cycles_between_CIP) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"
            if len(list_of_tuples) > 0:
                cursor.executemany(sql_statement, list_of_tuples)

            print("13th Sheet")
            rsheet13 = pd.DataFrame(columns=head13)
            df_collection[15] = bbt2(rsheet13,  ws1,
                                     zone, country, city1, brewery, dte, cleaned_inputblob_name)
            df_collection[15].columns = head13
            list_of_tuples = [
                tuple(x) for x in df_collection[15].to_records(index=False)]
            cursor.fast_executemany = True
            sql_statement = "INSERT INTO BBT_Cellars ( Filename , Zone , Country , City , Brewery , Date , Tank_Group , Tanks_nbr , Gross_Volume_per_Tank_hl , Main_Supplier , Year , Type , Average_Filling_Rate_percent , Net_Volume_hl , Average_Filling_Time_hours , Minimum_Residence_Time_hours , Tank_Group_used_for_small_batches , Additional_time_for_small_batch_multiple_release_hours , Average_Emptying_Time_hours , Accepted_Downtime_hours , Phasing_hours , Scheduling_hours , Manual_connections_hours , Other_please_specify_in_free_text_area_below , Average_Cycle_Time_hours_excluding_CIP , Real_CIP_Time_hours_per_CIP , Average_CIP_Time_hours_lost_per_cycle , Average_Cycle_Time_Including_CIP_hours , Average_nbr_Cycles_per_Week , Week_Available_Bright_Beer_Tank_Capacity_hl) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"
            if len(list_of_tuples) > 0:
                cursor.executemany(sql_statement, list_of_tuples)

            print("14th Sheet")
            ws1 = s1.sheet_by_name("4.1 - GLASS BOTTLE LINES ")
            rsheet17 = pd.DataFrame(columns=head17)
            df_collection[16] = gbottle(rsheet17,  ws1,
                                        zone, country, city1, brewery, dte, cleaned_inputblob_name)
            df_collection[16].columns = head17
            list_of_tuples = [
                tuple(x) for x in df_collection[16].to_records(index=False)]
            cursor.fast_executemany = True
            sql_statement = "INSERT INTO Glass_Lines ( Filename , Zone , Country , City , Brewery , Date , Line , Line_Name , SKU_Container_Type , Bottle_Content_liters , Bottle_Nominal_Speed_Bph , Can_Week_Maximum_Capacity_hl , Can_directly_to_Transport_pack_or_pallet , SKU_secondary_Pack_Type , Content_Bottles_per_secondary_pack , Content_hl_per_secondary_pack , Calculated_Nominal_Speed_secondary_packs_per_hrs , Forced_Nominal_Speed_secondary_packs_per_hrs , secondary_pack_Week_Maximum_Capacity_hl , SKU_Transport_Pack_Type , Content_Bottle_per_transport_pack , Content_hl_per_transport_pack , Calculated_Nominal_Speed_Transport_packs_per_hrs , Forced_Nominal_Speed_Transport_packs_per_hrs , Transport_pack_Week_Maximum_Capacity_hl , Pallets_Packs_per_layer , Pallet_nbr_layers , Packs_per_pallet , hl_per_pallet , Calculated_Nominal_Speed_pallets_per_hrs , Forced_Nominal_Speed_pallets_per_hrs , Pallets_Week_Maximum_Capacity_hl , Lowest_Week_Maximum_Capacity_hl , LEF_percent , Week_GLY_OAE_percent , Calculated_MIX_Capacity , Forced_MIX_Capacity , Time_to_Schedule_hrs_per_year , Week_Engineering_Capacity_khl , Month_Engineering_Capacity_khl , Year_Engineering_Capacity_khl , Sales_Peak_avg_3_months_percent , Seasonal_factor , Year_Operational_Capacity_khl , Total_Budgeted_Volumes_per_SKU_type_SKU_MIX_percent , Year_Total_Budgeted_Volumes_per_SKU_type , P_month_Total_Budgeted_Volumes_per_SKU_type) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"
            if len(list_of_tuples) > 0:
                cursor.executemany(sql_statement, list_of_tuples)

            print("15th Sheet")
            ws1 = s1.sheet_by_name("4.2 - CAN LINES")
            rsheet18 = pd.DataFrame(columns=head18)
            df_collection[17] = canl(rsheet18,  ws1,
                                     zone, country, city1, brewery, dte, cleaned_inputblob_name)
            df_collection[17].columns = head18
            list_of_tuples = [
                tuple(x) for x in df_collection[17].to_records(index=False)]
            cursor.fast_executemany = True
            sql_statement = "INSERT INTO Can_Lines (Filename , Zone , Country , City , Brewery , Date , Line , Line_Name , SKU_Container_Type , Can_Content_liters , Can_Nominal_Speed_Bph , Can_Week_Maximum_Capacity_hl , Can_directly_to_Transport_pack_or_pallet , SKU_secondary_Pack_Type , Content_Cans_per_secondary_pack , Content_hl_per_secondary_pack , Calculated_Nominal_Speed_secondary_packs_per_hrs , Forced_Nominal_Speed_secondary_packs_per_hrs , secondary_pack_Week_Maximum_Capacity_hl , SKU_Transport_Pack_Type , Content_Can_per_transport_pack , Content_hl_per_transport_pack , Calculated_Nominal_Speed_Transport_packs_per_hrs , Forced_Nominal_Speed_Transport_packs_per_hrs , Transport_pack_Week_Maximum_Capacity_hl , Pallets_Cans_per_layer , Pallet_nbr_layers , Cans_per_pallet , hl_per_pallet , Calculated_Nominal_Speed_pallets_per_hrs , Forced_Nominal_Speed_pallets_per_hrs , Pallets_Week_Maximum_Capacity_hl , Lowest_Week_Maximum_Capacity_hl , LEF_percent , Week_GLY_OAE_percent , Calculated_MIX_Capacity , Forced_MIX_Capacity , Time_to_Schedule_hrs_per_year , Week_Engineering_Capacity_khl , Month_Engineering_Capacity_khl , Year_Engineering_Capacity_khl , Sales_Peak_avg_3_months_percent , Seasonal_factor , Year_Operational_Capacity_khl , Total_Budgeted_Volumes_per_SKU_type_SKU_MIX_percent , Year_Total_Budgeted_Volumes_per_SKU_type , P_month_Total_Budgeted_Volumes_per_SKU_type ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"
            if len(list_of_tuples) > 0:
                cursor.executemany(sql_statement, list_of_tuples)

            print("16th Sheet")
            ws1 = s1.sheet_by_name("4.3 - PET LINES")
            rsheet19 = pd.DataFrame(columns=head19)
            df_collection[18] = petl(rsheet19,  ws1,
                                     zone, country, city1, brewery, dte, cleaned_inputblob_name)
            df_collection[18].columns = head19
            list_of_tuples = [
                tuple(x) for x in df_collection[18].to_records(index=False)]
            cursor.fast_executemany = True
            sql_statement = "INSERT INTO PET_lines (Filename , Zone , Country , City , Brewery , Date , Line , Line_Name , SKU_Container_Type , PET_Content_liters , PET_Nominal_Speed_Bph , PET_Week_Maximum_Capacity_hl , PET_directly_to_Transport_pack_or_pallet , SKU_secondary_Pack_Type , Content_PETs_per_secondary_pack , Content_hl_per_secondary_pack , Calculated_Nominal_Speed_secondary_packs_per_hrs , Forced_Nominal_Speed_secondary_packs_per_hrs , secondary_pack_Week_Maximum_Capacity_hl , SKU_Transport_Pack_Type , Content_PET_per_transport_pack , Content_hl_per_transport_pack , Calculated_Nominal_Speed_Transport_packs_per_hrs , Forced_Nominal_Speed_Transport_packs_per_hrs , Transport_pack_Week_Maximum_Capacity_hl , Pallets_PETs_per_layer , Pallet_nbr_layers , PETs_per_pallet , hl_per_pallet , Calculated_Nominal_Speed_pallets_per_hrs , Forced_Nominal_Speed_pallets_per_hrs , Pallets_Week_Maximum_Capacity_hl , Lowest_Week_Maximum_Capacity_hl , LEF_percent , Week_GLY_OAE_percent , Calculated_MIX_Capacity , Forced_MIX_Capacity , Time_to_Schedule_hrs_per_year , Week_Engineering_Capacity_khl , Month_Engineering_Capacity_khl , Year_Engineering_Capacity_khl , Sales_Peak_avg_3_months_percent , Seasonal_factor , Year_Operational_Capacity_khl , Total_Budgeted_Volumes_per_SKU_type_SKU_MIX_percent , Year_Total_Budgeted_Volumes_per_SKU_type , P_month_Total_Budgeted_Volumes_per_SKU_type) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"
            if len(list_of_tuples) > 0:
                cursor.executemany(sql_statement, list_of_tuples)

            print("17th Sheet")
            ws1 = s1.sheet_by_name("4.4 - KEG LINES")
            rsheet20 = pd.DataFrame(columns=head20)
            df_collection[19] = keg1(rsheet20,  ws1,
                                     zone, country, city1, brewery, dte, cleaned_inputblob_name)
            df_collection[19].columns = head20
            list_of_tuples = [
                tuple(x) for x in df_collection[19].to_records(index=False)]
            cursor.fast_executemany = True
            sql_statement = "INSERT INTO KEG_lines ( Filename , Zone , Country , City , Brewery , Date , Line , SKU_Keg_Type , Keg_Content_liters , Keg_Nominal_Speed_Bph , Keg_Week_Maximum_Capacity_hl , Keg_directly_to_Transport_pack_or_pallet , SKU_secondary_Pack_Type , Content_Kegs_per_secondary_pack , Content_hl_per_secondary_pack , Calculated_Nominal_Speed_secondary_packs_per_hrs , Forced_Nominal_Speed_secondary_packs_per_hrs , secondary_pack_Week_Maximum_Capacity_hl , Kegs_or_secondary_pack_directly_to__pallet , SKU_Transport_Pack_Type , Content_Kegs_per_transport_pack , Content_hl_per_transport_pack , Calculated_Nominal_Speed_Transport_packs_per_hrs , Forced_Nominal_Speed_Transport_packs_per_hrs , Transport_pack_Week_Maximum_Capacity_hl , Pallets_Kegs_per_layer , Pallet_nbr_layers , Kegs_per_pallet , hl_per_pallet , Calculated_Nominal_Speed_pallets_per_hrs , Forced_Nominal_Speed_pallets_per_hrs , Pallets_Week_Maximum_Capacity_hl , Lowest_Week_Maximum_Capacity_hl , LEF_percent , Week_GLY_OAE_percent , Calculated_MIX_Capacity , Forced_MIX_Capacity , Time_to_Schedule_hrs_per_year , Week_Engineering_Capacity_khl , Month_Engineering_Capacity_khl , Year_Engineering_Capacity_khl , Sales_Peak_avg_3_months_percent , Seasonal_factor , Year_Operational_Capacity_khl , Total_Budgeted_Volumes_per_SKU_type_SKU_MIX_percent , Year_Total_Budgeted_Volumes_per_SKU_type , P_month_Total_Budgeted_Volumes_per_SKU_type) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"
            if len(list_of_tuples) > 0:
                cursor.executemany(sql_statement, list_of_tuples)

            print("18th Sheet")
            ws1 = s1.sheet_by_name("4.5 - SPECIAL KEG LINES")
            rsheet21 = pd.DataFrame(columns=head21)
            df_collection[20] = skeg1(rsheet21,  ws1,
                                      zone, country, city1, brewery, dte, cleaned_inputblob_name)
            df_collection[20].columns = head21
            list_of_tuples = [
                tuple(x) for x in df_collection[20].to_records(index=False)]
            cursor.fast_executemany = True
            sql_statement = "INSERT INTO Special_KEG_lines ( Filename , Zone , Country , City , Brewery , Date , Line , SKU_Keg_Type , Keg_Content_liters , Keg_Nominal_Speed_Bph , Keg_Week_Maximum_Capacity_hl , Keg_directly_to_Transport_pack_or_pallet , SKU_secondary_Pack_Type , Content_Kegs_per_secondary_pack , Content_hl_per_secondary_pack , Calculated_Nominal_Speed_secondary_packs_per_hrs , Forced_Nominal_Speed_secondary_packs_per_hrs , secondary_pack_Week_Maximum_Capacity_hl , Kegs_or_secondary_pack_directly_to__pallet , SKU_Transport_Pack_Type , Content_Kegs_per_transport_pack , Content_hl_per_transport_pack , Calculated_Nominal_Speed_Transport_packs_per_hrs , Forced_Nominal_Speed_Transport_packs_per_hrs , Transport_pack_Week_Maximum_Capacity_hl , Pallets_Kegs_per_layer , Pallet_nbr_layers , Kegs_per_pallet , hl_per_pallet , Calculated_Nominal_Speed_pallets_per_hrs , Forced_Nominal_Speed_pallets_per_hrs , Pallets_Week_Maximum_Capacity_hl , Lowest_Week_Maximum_Capacity_hl , LEF_percent , Week_GLY_OAE_percent , Calculated_MIX_Capacity , Forced_MIX_Capacity , Time_to_Schedule_hrs_per_year , Week_Engineering_Capacity_khl , Month_Engineering_Capacity_khl , Year_Engineering_Capacity_khl , Sales_Peak_avg_3_months_percent , Seasonal_factor , Year_Operational_Capacity_khl , Total_Budgeted_Volumes_per_SKU_type_SKU_MIX_percent , Year_Total_Budgeted_Volumes_per_SKU_type , P_month_Total_Budgeted_Volumes_per_SKU_type) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) "
            if len(list_of_tuples) > 0:
                cursor.executemany(sql_statement, list_of_tuples)

            print("19th Sheet")
            rsheet22 = pd.DataFrame(columns=head22)
            df_collection[21] = pdown(
                rsheet22, cleaned_inputblob_name, s1, zone, country, city1, brewery, dte)
            df_collection[21].columns = head22
            list_of_tuples = [
                tuple(x) for x in df_collection[21].to_records(index=False)]
            cursor.fast_executemany = True
            sql_statement = "INSERT INTO Pack_Downtime ( Filename , Zone , Country , City , Brewery , Date , Line , Glass , Week_Total_Time , Week_Non_Schedule_Time , Week_Schedule_Time , Downtime_for_Yearly_Overhaul) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)"
            if len(list_of_tuples) > 0:
                cursor.executemany(sql_statement, list_of_tuples)

            logging.info("file uploaded complete")